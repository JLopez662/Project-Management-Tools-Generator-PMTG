import pandas as pd
import re
from datetime import timedelta, datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Color, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

def allocateTasksToWeeks(tasks):
    colWeekHours = [40] #first column, representing a work week, or 40 hours
    chronogram = [] 

    for task in tasks:
        weeks = len(colWeekHours)
        taskRow = ['_'] * weeks #current row times the weeks needed 
        while task > 0: #While the task has hours left to assing
            for i in range(len(colWeekHours)):
                if task <= colWeekHours[i]: #task needs less hours than available in current work week
                    colWeekHours[i] -= task
                    taskRow[i] = 'X'
                    task = 0 #Task hours fully allocated
                    break
                else:        #Task needs more hours than available in current work week
                    if colWeekHours[i] > 0:
                        task -= colWeekHours[i]
                        taskRow[i] = 'X'
                        colWeekHours[i] = 0

            #If task still has hours left not allocated, add new week
            if task > 0:   
                colWeekHours.append(40)
                taskRow.append('_')    #Extend task row for the new week

        #Update chronogram with next task
        chronogram.append(taskRow) 

         #Add weeks not used by tasks 
        for row in chronogram: 
            while len(row) < len(taskRow):
                row.append('_')

    return chronogram

# Function to validate the date format MM/DD
def validate_date(date_text):
    try:
        datetime.strptime(date_text, '%m/%d')
        return True
    except ValueError:
        return False
    
# Function to calculate date ranges for each week, spanning 7 days each
def get_week_dates(start_date, num_weeks, year):
    week_dates = []
    start_date = datetime.strptime(f"{start_date}/{year}", "%m/%d/%Y")
    for i in range(num_weeks + 1):
        end_date = start_date + timedelta(days=6)
        # Format the date as '01/Feb - 07/Feb'
        week_dates.append(f'{start_date.strftime("%d/%b")} - {end_date.strftime("%d/%b")}')
        start_date = end_date + timedelta(days=1)
    return week_dates
    

def chronogramToExcel(chronogram, year, start_week, filename="chronogram.xlsx"):
    #Create DataFrame from chronogram
    df = pd.DataFrame(chronogram)
    df.insert(0, ' ', '')
    
    #Write DataFrame to Excel file
    df.to_excel(filename, index=False, header=False)

    #If year is not provided, use current year
    if not year:
        year = datetime.now().year
    else:
        try:
            year = int(year)
        except:
            print("The year format is incorrect. Please use YYYY format for the year.")
            return

    # Assuming that the month headers start from the second column (B)
    # and span to the last column with data.
    last_data_column = len(df.columns) + 1# Plus one because dataframe indexing starts at 0

    #Open Excel file and color cells
    wb = Workbook()
    ws = wb.active #Get current active sheet

    # Merge cells for the year header from the second column to the last data column
    ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=last_data_column)

    # Set the value, style, and alignment for the year header
    year_cell = ws.cell(row=1, column=2)
    year_cell.value = str(year)
    year_cell.alignment = Alignment(horizontal='center', vertical='center')
    year_cell.fill = PatternFill(start_color="0070c0", end_color="0070c0", fill_type="solid")
    year_cell.font = Font(color="FFFFFF")

    # Insert month headers aligned with the week date ranges
    week_dates = get_week_dates(start_week, last_data_column - 2, year)  # -2 accounts for the index and the initial space
    months = {}
    for i, week_date in enumerate(week_dates, start=2):  # start=2 to account for the initial empty column
        month_name = datetime.strptime(week_date.split(' - ')[0], "%d/%b").strftime("%B")
        if month_name not in months:
            months[month_name] = {'start': i, 'end': i}
        else:
            months[month_name]['end'] = i

    # Add month headers starting from the second row
    row_offset = 2
    for month, cols in sorted(months.items()):
        ws.merge_cells(start_row=2, start_column=cols['start'], end_row=2, end_column=cols['end'])
        month_cell = ws.cell(row=row_offset, column=cols['start'])
        month_cell.value = month
        #month_cell.alignment = Alignment(horizontal='center')
        month_cell.fill = PatternFill(start_color="0070c0", end_color="0070c0", fill_type="solid")
        month_cell.font = Font(color="FFFFFF")

    # Adjust the row offset to start adding week dates
    row_offset += 1

    # Add week dates below month headers
    for col, date_range in enumerate(week_dates, start=2):  # Again, start=2 for the initial empty column
        week_cell = ws.cell(row=row_offset, column=col)
        week_cell.value = date_range
        week_cell.alignment = Alignment(horizontal='center')
        week_cell.fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
        week_cell.font = Font(color="FFFFFF")

    # Now adjust row_offset to start adding tasks below the week headers
    row_offset += 1

    # Add tasks to the Excel sheet
    for index, row in enumerate(chronogram, start=row_offset):
        for col_index, value in enumerate(row, start=2):  # start=2 for the initial empty column
            task_cell = ws.cell(row=index, column=col_index)
            if value == 'X':
                task_cell.fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")

    # Set column widths
    column_width = 15
    for col in ws.iter_cols(min_col=2, max_col=ws.max_column, min_row=1, max_row=ws.max_row):
        for cell in col:
            #cell.alignment = Alignment(wrap_text=True)
            ws.column_dimensions[get_column_letter(cell.column)].width = column_width

    
    wb.save(filename)
    df.to_csv("chronogram.csv", index=False)  # Also save as CSV


##########
    #header = pd.DataFrame(columns=['2024'])

    #df = pd.concat([header, df], axis=1)
    #ws.cell(row=1, column=1).fill=PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
################

# Ask user for the year for the Gantt Chart
yearInput = input("Add the year for the Gantt Chart (leave empty if using current year): ")
if not yearInput.strip():
    yearInput = str(datetime.now().year)


# Prompt the user for the starting week, now expecting MM/DD format
start_week = input("Add the starting week (MM/DD) (leave empty if not): ")
while start_week and not validate_date(start_week):
    start_week = input("The format is incorrect. Please use MM/DD format or leave empty: ")


#Ask user for input (hours as separated values by comma)
taskHoursInput = input("Add tasks hours (as comma-separated values): ")

#Convert input string removing extra spaces and split input into a list of integers
tasks = [int(x.strip()) for x in re.split(r'[,\s]+', taskHoursInput) if x.strip()]

#Generate the chronogram from user input
chronogram = allocateTasksToWeeks(tasks)

#Save chronogram to Excel with colored cells format
chronogramToExcel(chronogram, yearInput, start_week, "chronogram.xlsx")




        

